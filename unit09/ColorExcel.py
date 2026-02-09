import openpyxl
from openpyxl.styles import PatternFill
import os

current_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(current_dir,"..", "etc", "./test.xlsx")

try:
    # 2. 파일 불러오기
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 3. 데이터 입력
    ws['B1'] = "is"
    ws['C1'] = "good"

    # 4. 단일 셀 색상 변경 (C1 -> 초록색)
    green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
    ws['C1'].fill = green_fill

    # 5. 범위 색상 변경 (A2:C2 -> 노란색)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # A2, B2, C2 각각 칠하기
    ws['A2'].fill = yellow_fill
    ws['B2'].fill = yellow_fill
    ws['C2'].fill = yellow_fill

# 6. 저장하기
    wb.save(file_path)
    print("색칠 완료: C1(초록), A2:C2(노랑) 색상이 적용되었습니다!")

except FileNotFoundError:
    print("error: 'test.xlsx' 파일이 없습니다. CreateExcel.py를 먼저 실행해주세요.")