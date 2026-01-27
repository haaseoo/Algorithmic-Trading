import openpyxl
import os

currentDir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(currentDir, "..", "etc", "test.xlsx")

print(f"읽어올 파일 경로 : {file_path}")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    val1 = ws.cell(row=1, column=1).value
    print(f"A1 값 (좌표): {val1}")

    val2 = ws["A1"].value
    print(f"A1셀 값 (주소): {val2}")

    wb.close()

except FileNotFoundError:
    print("파일을 찾을 수 없습니다.")